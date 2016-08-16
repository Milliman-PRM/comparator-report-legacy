"""
### CODE OWNERS: Jason Altieri, Aaron Burgess

### OBJECTIVE:
  Extract contents of various excel files for later use

### DEVELOPER NOTES:
  Requires a full PRM environment when executed as a script.
"""
import csv
import logging
import re
import shutil
import string
from collections import defaultdict, Counter, OrderedDict, namedtuple
from datetime import datetime
from openpyxl import load_workbook
import prm.meta.project
from indypy.file_utils import IndyPyPath

PRM_META = prm.meta.project.parse_project_metadata()

LOGGER = logging.getLogger(__name__)

_QUARTER_RANGE_DICT = {
    1: 4,
    2: 4,
    3: 4,
    4: 1,
    5: 1,
    6: 1,
    7: 2,
    8: 2,
    9: 2,
    10: 3,
    11: 3,
    12: 3
}

_QUARTER_DATE_RANGE_MAP = {
    1: {'start': '1-1-{}', 'end_date': '3-31-{}'},
    2: {'start': '4-1-{}', 'end_date': '6-30-{}'},
    3: {'start': '7-1-{}', 'end_date': '9-30-{}'},
    4: {'start': '10-1-{}', 'end_date': '12-31-{}'}
}

_FIELD_NAMES = ['HICNO', 'First Name', 'Last Name']
_INCLUDE_TABS = ['TABLE 1-1', 'TABLE 1-2', 'TABLE 1-3', 'TABLE 1-4', 'TABLE 1-5', 'TABLE 1-6']
_EXCLUDE_TABS = ['COVER', 'TOC', 'GLOSSARY', 'PARAMETERS']
_TABLE5_KEYWORD = 'HICNOs of beneficiaries'
_PROSPECTIVE_KEYWORD = 'Prospective'
_LIMIT_QASSIGN_PATTERN = 'Quarter'

_TABLE_FIELD_DICT = OrderedDict([
    ('TABLE 1-4', _FIELD_NAMES + ['ACO Participant TIN', 'Individual NPI']),
    ('TABLE 1-3', _FIELD_NAMES + ['ACO Participant CCN Number']),
    ('TABLE 1-2', _FIELD_NAMES + ['ACO Participant TIN Number']),
    ('TABLE 1-5', _FIELD_NAMES + ['Date of Death', 'NotAssigned1', 'NotAssigned2', 'NotAssigned3',
                                  'NotAssigned4', 'NotAssigned5', 'NotAssigned6']),
    ('TABLE 1-6', _FIELD_NAMES + ['Date of Death']),
    ('TABLE 1-1', _FIELD_NAMES + ['Monthly Eligibility Flag {}'.format(x) for x in range(1, 13)])
])

_FILE_DATE_PATTERN_REGEX = re.compile(r'(QASSGN|HASSGN).+?(D\d{6}\.T\d{6})')
_QASSGN_DATE_PATTERN = re.compile(r'(Year \d{4}(?:.+?)Quarter \d)', re.I)
_HASSGN_DATE_PATTERN = re.compile(r'(Year \d{4})(?!, Quarter)', re.I)


def _exclude_files(directory, anti_pattern):
    """Search for files we want, files we don't and files that are ambiguous"""
    main_directory = IndyPyPath(directory)
    anti_files = main_directory.collect_files_regex(anti_pattern)
    potential_pro_files = [file for file in main_directory.rglob('*.*')
                           if file not in anti_files]
    convert_files = [file for file in
                     potential_pro_files if re.search(r'T\d{6}', str(file))]
    for file in convert_files:
        if re.search(r'T\d{6}$', str(file)):
            new_file_str = str(file) + '.xlsx'
            shutil.move(str(file), new_file_str)
        if re.search(r'T\d{6}\.xls$', str(file)):  # default format is .xlsx; .xls is likely a typo
            new_file_str = str(file) + '.xlsx'
            shutil.copy(str(file), new_file_str)
    return [final_file for final_file
            in main_directory.collect_files_extensions(['xlsx'])
            if final_file not in anti_files and str(final_file).find('~$') == -1]


def _create_date_pattern_dictionary(included_file_list):
    """Determine which files can deterministically be included by file name"""
    file_dictionary = defaultdict(list)
    for file in included_file_list:
        date_pattern_check = _FILE_DATE_PATTERN_REGEX.findall(str(file))
        if not date_pattern_check:
            file_dictionary['UNKNOWN'].append(file)
            continue
        date_pattern_results = date_pattern_check[0]
        if len(date_pattern_results) == 1:
            file_dictionary['UNKNOWN'].append(file)
        else:
            year_quarter_key = _extract_date_from_name(date_pattern_results[1])
            file_dictionary[year_quarter_key].append(file)
    return file_dictionary


def _scrape_tab_names(workbook_list):
    """"Get a list of tab names in each workbook"""
    sheet_dictionary = {}
    for workbook in map(str, workbook_list):
        wb = load_workbook(filename=workbook)
        sheet_dictionary[str(workbook)] = wb.get_sheet_names()
    return sheet_dictionary


def _identify_useful_tabs(sheet_dictionary):
    """Iterate over workbook: sheet tabs dictionary and assign usefulness of
    tab names"""
    file_useful_tabs_dict = {}
    for file_path, tabs in sheet_dictionary.items():
        assigned_files_dict = {'INCLUDE': [], 'EXCLUDE': [], 'UNKNOWN': []}
        for tab in tabs:
            if tab.upper() in _INCLUDE_TABS:
                assigned_files_dict['INCLUDE'].append(tab)
            elif tab.upper() in _EXCLUDE_TABS:
                assigned_files_dict['EXCLUDE'].append(tab)
            else:
                assigned_files_dict['UNKNOWN'].append(tab)
        file_useful_tabs_dict[file_path] = assigned_files_dict
    return file_useful_tabs_dict
        

def _extract_date_from_name(date_pattern):
    """Take date string from file name and create year.quarter key"""
    file_date = datetime.strptime('20' + date_pattern[1:5], '%Y%m')
    quarter = _QUARTER_RANGE_DICT[file_date.month]
    year = file_date.year if quarter != 4 else int(file_date.year) - 1
    return "{year}.{quarter}".format(year=year, quarter=quarter)


def _scan_sheets_for_date(worksheet):
    """Scan the worksheets for year and/or quarter indicators"""
    rows = worksheet.iter_rows('A1:L20')
    date_counter = Counter()
    for row in rows:
        for cell in row:
            if cell.value:
                qassgn_match = [_extract_date_from_sheet(result)
                            for result in _QASSGN_DATE_PATTERN.findall(str(cell.value))]
                if qassgn_match:
                    date_counter.update(qassgn_match)
                else:
                    hassgn_match = [_extract_date_from_sheet(result)
                                    for result in _HASSGN_DATE_PATTERN.findall(str(cell.value))]
                    if hassgn_match:
                        date_counter.update(hassgn_match)
    return date_counter


def _extract_date_from_sheet(date_match):
    """Return a 'YYYY.Q' key from a match in the sheet text"""
    numeric_only = re.sub(r'[^0-9]+', '', date_match)
    if len(numeric_only) == 4:
        return numeric_only
    else:
        return numeric_only[:4] + '.' + numeric_only[-1]


def _check_for_field_names(worksheet):
    """Iterate over unknown worksheets if missing key tables"""
    rows = worksheet.iter_rows('A1:N20')

    def _keyword_check(values, keyword):   # pragma: no cover
        """Check if dealing with table 5"""
        row_string = ' '.join(values)
        if re.match(keyword, row_string):
            return True

    prospective_check = False
    for row_number, row in enumerate(rows):
        row_values = [str(cell.value).upper().strip() for cell in row if cell.value]
        if not row_values:
            continue
        if worksheet.title.upper().strip() in _TABLE_FIELD_DICT:
            if _keyword_check(row_values, _PROSPECTIVE_KEYWORD):
                prospective_check = True
                if _keyword_check(row_values, _LIMIT_QASSIGN_PATTERN):
                    prospective_check = False
            if worksheet.title.upper().strip() == 'TABLE 1-5':
                if _keyword_check(row_values, _TABLE5_KEYWORD):
                    return worksheet.title, row_number, prospective_check
            upper_case_fields = [
                field.upper().strip()
                for field in _TABLE_FIELD_DICT[worksheet.title.upper().strip()]
                if field.upper().find('ELIGIBILITY') == -1  # may not be present
                ]
            field_check = set(upper_case_fields) <= set(row_values)
            if field_check:
                return worksheet.title, row_number, prospective_check
        else:
            for table, fields in _TABLE_FIELD_DICT.items():
                if _keyword_check(row_values, _PROSPECTIVE_KEYWORD):
                    prospective_check = True
                    if _keyword_check(row_values, _LIMIT_QASSIGN_PATTERN):
                        prospective_check = False
                upper_case_fields = [
                    field.upper().strip() for field in fields
                    if field.upper().find('ELIGIBILITY') == -1  # may not be present
                    ]
                field_check = set(upper_case_fields) <= set(row_values)
                if field_check:
                    return table, row_number, prospective_check


def _build_final_dict(annotated_tab_dict):
    """Build dict with field path key to list of tuples with included table name
    and row position of beginning field names"""
    final_dict = defaultdict(list)
    valuable_tab_nt = namedtuple('valuable_tab_nt', ['actual_tab_name',
                                                     'inferred_table_name',
                                                     'header_row',
                                                     'prospective_flag'])
    for table_path, inclusion_dict in annotated_tab_dict.items():
        wb = load_workbook(table_path)
        for worksheet in wb.worksheets:
            if worksheet.title in inclusion_dict['EXCLUDE']:
                continue
            check_for_table = _check_for_field_names(worksheet)
            if check_for_table:
                if check_for_table[0] not in final_dict[table_path]:
                    tab_info = valuable_tab_nt(worksheet.title, *check_for_table)
                    final_dict[table_path].append(tab_info)
    return final_dict


def _create_empty_tables():   # pragma: no cover
    """Creates empty csvs with headers ready for use later"""
    for table, fields in _TABLE_FIELD_DICT.items():
        final_table_name = 'table_{}.csv'.format(table[-1])
        headers = ['date_start', 'date_end'] + fields + ['hassgn']
        table_path = IndyPyPath(_TARGET_PATH) / final_table_name
        with open(str(table_path), 'w', newline='') as outfile:
            writer = csv.DictWriter(outfile, headers)
            writer.writeheader()


def _map_start_end_date(inferred_date):
    """Determine start and end date for data given the inferrend date"""
    if len(inferred_date) == 4:
        date_start = '1-1-{}'.format(inferred_date)
        date_end = '12-31-{}'.format(inferred_date)
    else:
        date_dict = _QUARTER_DATE_RANGE_MAP[int(inferred_date[-1])]
        date_start = date_dict['start'].format(inferred_date.split('.')[0])
        date_end = date_dict['end_date'].format(inferred_date.split('.')[0])
    return date_start, date_end


def _write_data_to_csvs(final_dictionary):
    """Iterate over final dictionary and write tables out, appending
    by occurrence, gather date from worksheets"""
    for workbook, mapped_tabs_list in final_dictionary.items():
        total_counter = Counter()
        wb = load_workbook(workbook)
        for worksheet in wb.worksheets:
            total_counter.update(_scan_sheets_for_date(worksheet))
        try:
            actual_date = total_counter.most_common(1)[0][0]
        except IndexError:
            LOGGER.debug("Actual Date not found for file %s", str(workbook))
            continue
        date_range = _map_start_end_date(actual_date)
        if date_range[0].find('1-1') > -1 and date_range[1].find('12-31') > -1:
            hassgn = "TRUE"
        else:
            hassgn = "FALSE"
        for values in mapped_tabs_list:
            table5_indicator = False
            table5_count = 0
            if values.prospective_flag:
                hassgn = "PROSP"
            final_table_name = 'table_{}.csv'.format(values.inferred_table_name[-1])
            write_path = IndyPyPath(_TARGET_PATH) / final_table_name
            with open(str(write_path), 'a', newline='') as outfile:
                writer = csv.writer(outfile)
                header_columns = []
                for row_number, row in enumerate(wb[values.actual_tab_name].iter_rows()):
                    if row_number < values.header_row:
                        continue
                    elif row_number == values.header_row:
                        headers = [name.upper().strip()
                                   for name in _TABLE_FIELD_DICT[values.inferred_table_name.upper()]
                                   ]
                        if not values.inferred_table_name[-1] == '5':
                            for cell in row:
                                clean_cell_value = re.sub(  # remove duplicate spaces
                                    ' +',
                                    ' ',
                                    str(cell.value).upper().strip()
                                )
                                if clean_cell_value.find("ELIGIBILITY") > -1:
                                    clean_cell_value = clean_cell_value[:-1]  # remove superscript
                                if clean_cell_value in headers:
                                    header_columns.append(cell.column)
                        else:
                            header_columns = _table_five_column_inference(row)
                            table5_indicator = True
                        continue
                    elif table5_indicator:
                        table5_count += 1
                        if table5_count == 2:
                            table5_indicator = False
                        continue
                    else:
                        clean_row = [str(cell.value).strip()
                                     if cell.value is not None else ''
                                     for cell in row
                                     if cell.column in header_columns]
                        if list(set(clean_row)) == [''] and row_number > values.header_row + 1:
                            break
                        count_missing = len(headers) - len(clean_row)
                        for n in range(count_missing):
                            clean_row.append('')  # append blanks when elig status flag not present
                        final_row = list(date_range) + clean_row + [hassgn]
                        writer.writerow(final_row)


def _table_five_column_inference(row):
    """Identify known values for headers and infer additional header positions"""
    header_columns = []
    for cell in row:
        if str(cell.value).upper().strip().find(_TABLE5_KEYWORD.upper()) > -1:
            header_columns.append(cell.column)
        elif str(cell.value).upper().strip() in ['FIRST NAME', 'LAST NAME']:
            header_columns.append(cell.column)
        elif re.sub(r'[^A-Z ]+', '', str(cell.value).upper().strip()) \
                in ['DECEASED BENEFICIARY FLAG', 'DATE OF DEATH']:
            header_columns.append(cell.column)
            column_index = string.ascii_uppercase.index(cell.column)
            for x in range(1, 7):
                header_columns.append(string.ascii_uppercase[column_index + x])
    return header_columns


def main():  # pragma: no cover
    """
    Crawl through reference files, detect assignment files, and parse out their useful
    contents into a format that is easier to consume
    """
    files = _exclude_files(IndyPyPath(_PATH_RECEIVED), 'Address')
    LOGGER.debug("%d files found", len(files))
    tab_name_dict = _scrape_tab_names(files)
    LOGGER.debug("Processed Tab Names to Workbook mappings...")
    annotated_tab_name_dict = _identify_useful_tabs(tab_name_dict)
    LOGGER.debug("Processed Tab Name Annotations")
    final_dict = _build_final_dict(annotated_tab_name_dict)
    LOGGER.debug("Final file scans for appropriate sheets")
    _create_empty_tables()
    LOGGER.debug("Writing data...")
    _write_data_to_csvs(final_dict)


if __name__ == '__main__':
    _TARGET_PATH = PRM_META[(17, 'out')]
    _PATH_RECEIVED = PRM_META["path_project_received_ref"]
    main()
