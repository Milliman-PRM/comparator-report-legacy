"""
### CODE OWNERS: Shea Parkes, Kyle Baird

### OBJECTIVE:
  Extract contents of various excel files for later use

### DEVELOPER NOTES:
  Requires a full PRM environment when executed as a script.
"""
import shutil
import re
import datetime

from datetime import date
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

# =============================================================================
# LIBRARIES, LOCATIONS, LITERALS, ETC. GO ABOVE HERE
# =============================================================================


class AssignmentWorksheet(object):
    """A single sheet in an MSSP assignment Workbook"""

    extract_fields = [
        'date_start',
        'date_end',
        'hicno',
        'tin',
        'npi',
        ]

    def __init__(self, ws):
        self.ws_obj = ws
        self.key_cells = dict()
        self.header_row = None
        self.date_start = None
        self.date_end = None

        self._sniff_header()
        self._calc_intrinsic_value()

    def _sniff_header(self):
        """Inspect the header for interesting information"""
        for row in self.ws_obj.get_squared_range(0, 0, 24, 16):
            for cell in row:
                if cell.value is None:
                    continue
                if not isinstance(cell.value, str):
                    continue
                if cell.value.lower() == 'hicno':
                    self.key_cells['hicno'] = cell
                    continue
                if re.search(r'\bparticipant tin\b', cell.value, re.I):
                    self.key_cells['tin'] = cell
                    continue
                if re.search(r'\bnpi\b', cell.value, re.I):
                    self.key_cells['npi'] = cell
                    continue
                if cell.value.lower() == 'year 2013 (october 2012-september 2013)':
                    self.date_start = date(2012, 10, 1)
                    self.date_end = date(2013, 9, 30)
                    continue
                match_quarters = re.search(
                    r'year\D*(?P<year>\d{4}).*Q\D*(?P<quarter>\d)',
                    cell.value,
                    re.IGNORECASE,
                    )
                if match_quarters:
                    self.date_end = date(
                        int(match_quarters.group('year')),
                        int(match_quarters.group('quarter')) * 3,
                        1,
                        ) # Not quite right yet, adjustments follow
                    self.date_start = (
                        self.date_end + datetime.timedelta(days=-45)
                        ).replace(day=1)
                    self.date_end = (
                        self.date_end + datetime.timedelta(days=45)
                        ).replace(day=1) + datetime.timedelta(days=-1)
                    continue
                match_annual = re.search(
                    r'year (?P<year>\d{4})',
                    cell.value,
                    re.IGNORECASE,
                    )
                if match_annual:
                    self.date_start = date(int(match_annual.group('year')), 1, 1)
                    self.date_end = date(int(match_annual.group('year')), 12, 31)
                    continue
        if self.date_start:
            assert self.date_start.day == 1, 'Windows must start on the first of the month'
            assert (self.date_end + datetime.timedelta(days=1)).day == 1, \
                'Windows must end on the last of the month'
            assert self.date_start.month % 3 == 1, 'Windows must start on a quarter'
            assert self.date_end.month % 3 == 0, 'Windows must end on a quarter'

    def _calc_intrinsic_value(self):
        """Calculate a score that represents the worth of this sheet"""
        interesting_rows = {cell.row for cell in self.key_cells.values()}

        if 'hicno' not in self.key_cells.keys():
            self.intrinsic_value = 0
        elif not self.date_end:
            self.intrinsic_value = 0
        elif len(interesting_rows) > 1:
            self.intrinsic_value = 0
        else:
            self.header_row = interesting_rows.pop()
            self.intrinsic_value = len(self.key_cells)

    def write_values(self, fh_out):
        """Write any discovered value"""
        assert self.intrinsic_value > 0, 'This worksheet is not worth extracting.'

        key_col_nums = {
            k: column_index_from_string(v.column) - 1
            for k, v
            in self.key_cells.items()
            }
        static_values = {
            'date_start': str(self.date_start),
            'date_end': str(self.date_end),
            }

        for row_num, row in enumerate(self.ws_obj.rows):
            if (row_num + 1) <= self.header_row:
                continue
            if row[key_col_nums['hicno']].value is None:
                break

            row_values = []
            for field in AssignmentWorksheet.extract_fields:
                try:
                    row_values.append(row[key_col_nums[field]].value)
                except KeyError:
                    try:
                        row_values.append(static_values[field])
                    except KeyError:
                        row_values.append('')
            fh_out.write('~'.join([value_.strip() for value_ in row_values]))
            fh_out.write('\n')

    @classmethod
    def write_header(cls, fh_out):
        """Write a header on anticipated output"""
        fh_out.write('~'.join(cls.extract_fields))
        fh_out.write('\n')


class AssignmentWorkbook(object):
    """An MSSP assignment workbook from CMS"""

    def __init__(self, wb_path):
        self.wb_path = wb_path
        self.wb_obj = load_workbook(
            str(wb_path),
            read_only=True,
            keep_vba=False,
            data_only=True,
            )
        self.worksheets = [AssignmentWorksheet(ws) for ws in self.wb_obj]
        self._max_intrinsic_value = max((
            ws.intrinsic_value
            for ws in self.worksheets
            ))
        self.key_worksheet = [
            ws
            for ws in self.worksheets
            if ws.intrinsic_value == self._max_intrinsic_value
            ][0]
        self.has_value = self._max_intrinsic_value > 0

    def write_values(self, fh_out):
        """Extract interesting values"""
        assert self.has_value, 'There is nothing worth extracting'
        self.key_worksheet.write_values(fh_out)

    @property
    def fields_found(self):
        """A quick introspection"""
        return self.key_worksheet.key_cells.keys()


def uncover_xlsx_files(path_sniffing):
    """Find potential assignment files that might be missing an 'xlsx' extension"""
    assert path_sniffing.is_dir(), '{} is not a directory'.format(str(path_sniffing))

    for file_ in path_sniffing.rglob('*'):
        if not file_.is_file():
            continue
        if file_.suffix.lower() == '.xlsx':
            continue
        if re.search(r'ACO\.[HQ]ASSGN\.D', file_.name, re.IGNORECASE):
            file_with_ext = file_.parent / (file_.name + '.xlsx')
            if file_with_ext.is_file():
                continue
            print('Duplicating {} with an "xlsx" extension\n\n'.format(file_))
            shutil.copy(str(file_), str(file_with_ext))


def main(path_root, filepath_out):
    """Process a directory of slop and put the clean info into an output file"""
    uncover_xlsx_files(path_root)
    with filepath_out.open('w') as fh_out:
        AssignmentWorksheet.write_header(fh_out)
        for path_wb in path_root.rglob('*.xlsx'):
            print('Processing {}\n'.format(path_wb.name))
            wb_ = AssignmentWorkbook(path_wb)
            print('Found the following fields: {}\n'.format(wb_.fields_found))
            print('Date range inferred to be from {} to {}\n\n'.format(
                wb_.key_worksheet.date_start,
                wb_.key_worksheet.date_end,
                ))
            if wb_.has_value:
                print('Workbook was determined to have value; data is being extracted.\n\n')
                wb_.write_values(fh_out)


if __name__ == '__main__':
    import sys
    import os
    sys.path.append(os.path.join(os.environ['USERPROFILE'], 'HealthBI_LocalData'))
    import healthbi_env
    if healthbi_env.META["project_id_prior"].lower() != "new":
        print("PULLING REFERENCES FROM PRIOR PROJECT\n\n")
        print([
            shutil.copy(str(file_), healthbi_env.META["path_project_received_ref"])
            for file_ in Path(
                healthbi_env.META["path_project_data_prior"] + "_From_Client" + os.sep
                ).rglob("*.xlsx")
            if file_.is_file()
            ], sep="\n")
        print("\n")
    print('BEGINNING TO SCAN RECEIVED DATA FOR ASSIGNMENT INFORMATION.\n\n')
    main(
        Path(healthbi_env.META['path_project_received']).parent,
        Path(healthbi_env.META[(17, 'out')]) / 'timeline_assign_extract.txt',
        )
    print('\n\nFINISHED SCANNING.\n')
